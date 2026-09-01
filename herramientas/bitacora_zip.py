# -*- coding: utf-8 -*-
"""
Relación de clientes por STATUS — MG Colima (versión módulo, sin Colab)

Refactor de Relacion_de_clientes_ZIP_DEL_MES.ipynb: toda la lógica de
extracción de PDFs y armado del Excel es EXACTAMENTE la misma que en el
notebook original; lo único que cambia es que en vez de subir el ZIP con
`files.upload()` (Colab), la función `procesar_zip_a_excel()` recibe la
ruta del ZIP como parámetro, para poder llamarse desde la app de Streamlit.

Requiere el binario `pdftotext` (paquete `poppler-utils` del sistema) —
ver packages.txt en la raíz del proyecto.
"""

import os
import re
import glob
import shutil
import zipfile
import subprocess
import unicodedata
import pandas as pd
from datetime import datetime
from pathlib import Path

# ─────────────────────────────────────────────
# Normalización de nombres de carpeta -> STATUS (plural/singular, variantes).
# Si aparece una carpeta nueva que no está aquí, se usa su nombre en mayúsculas
# tal cual, así que el notebook funciona aunque agreguen carpetas nuevas.
# ─────────────────────────────────────────────
MAPEO_STATUS = {
    "APROBADOS":        "APROBADO",
    "APROBADO":         "APROBADO",
    "RECHAZADOS":       "RECHAZADO",
    "RECHAZADO":        "RECHAZADO",
    "CONTRAPROPUESTAS": "CONTRAPROPUESTA",
    "CONTRAPROPUESTA":  "CONTRAPROPUESTA",
    "PENDIENTES":       "PENDIENTE",
    "PENDIENTE":        "PENDIENTE",
    "CANCELADOS":       "CANCELADO",
    "CANCELADO":        "CANCELADO",
    "EN PROCESO":       "EN PROCESO",
    "EN REVISION":      "EN REVISIÓN",
}

LIMPIAR_TEMPORAL = True   # borra los PDFs extraídos al terminar


def normalizar_status(nombre_carpeta: str) -> str:
    """Convierte el nombre de una carpeta en el valor de STATUS a usar."""
    nombre = nombre_carpeta.strip().upper()
    nombre = unicodedata.normalize("NFKD", nombre).encode("ascii", "ignore").decode("utf-8")
    return MAPEO_STATUS.get(nombre, nombre)


def extraer_zip(ruta_zip: str, carpeta_destino: str) -> str:
    """Extrae el ZIP y devuelve la ruta de la carpeta que contiene las
    subcarpetas de estatus (soporta ZIPs con una carpeta raíz envolvente,
    como 'JUNIO/Aprobados/...', y ZIPs sin ella)."""
    if os.path.isdir(carpeta_destino):
        shutil.rmtree(carpeta_destino)
    os.makedirs(carpeta_destino, exist_ok=True)

    with zipfile.ZipFile(ruta_zip, "r") as z:
        z.extractall(carpeta_destino)

    contenido = [
        os.path.join(carpeta_destino, n)
        for n in os.listdir(carpeta_destino)
        if not n.startswith("__MACOSX") and not n.startswith(".")
    ]
    subdirs = [c for c in contenido if os.path.isdir(c)]

    # Si el ZIP trae una única carpeta raíz envolviendo todo, se usa esa.
    if len(subdirs) == 1 and len(contenido) == 1:
        return subdirs[0]
    return carpeta_destino


def detectar_carpetas_status(carpeta_raiz: str) -> dict:
    """Detecta automáticamente las subcarpetas de estatus dentro de la carpeta raíz.
    Devuelve {nombre_carpeta_original: STATUS_normalizado}."""
    carpetas = {}
    for nombre in sorted(os.listdir(carpeta_raiz)):
        ruta = os.path.join(carpeta_raiz, nombre)
        if os.path.isdir(ruta) and not nombre.startswith(".") and not nombre.startswith("__MACOSX"):
            carpetas[nombre] = normalizar_status(nombre)
    return carpetas


def extraer_texto_pdf(ruta: str) -> str:
    resultado = subprocess.run(
        ["pdftotext", "-layout", "-enc", "UTF-8", ruta, "-"],
        capture_output=True
    )
    if resultado.returncode != 0:
        print(f"  ⚠ pdftotext falló (código {resultado.returncode}): {os.path.basename(ruta)}")
        return ""
    return resultado.stdout.decode("utf-8", errors="replace")


def limpiar(texto: str) -> str:
    return re.sub(r"\s+", " ", texto).strip()


def extraer_monto(cadena: str) -> str:
    match = re.search(r"\$\s*([\d,]+\.?\d*)", cadena)
    return f"${match.group(1)}" if match else cadena.strip()


def extraer_monto_float(cadena: str) -> float:
    """Convierte un monto tipo '$ 8,800.00' a float (0.0 si no hay monto)."""
    match = re.search(r"\$\s*([\d,]+\.?\d*)", cadena)
    return float(match.group(1).replace(",", "")) if match else 0.0


def normalizar_tipo_seguro(texto_tipo: str) -> str:
    """Normaliza el campo 'Tipo' de seguro a una de 3 categorías:
    MULTIANUAL FRACCIONADO, MULTIANUAL FINANCIADO o CONTADO ANUAL.
    Si aparece una variante nueva no reconocida, se devuelve el texto
    original en mayúsculas (así no se pierde información)."""
    t = texto_tipo.strip().upper()
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode("utf-8")
    if "FRACCIONADO" in t:
        return "MULTIANUAL FRACCIONADO"
    if "FINANCIADO" in t:
        return "MULTIANUAL FINANCIADO"
    if "CONTADO" in t and "ANUAL" in t:
        return "CONTADO ANUAL"
    return t


def parsear_solicitud(texto: str) -> dict:
    if not texto:
        return {}

    datos = {}

    m = re.search(r"Folio CCK:\s*(\d+)", texto)
    datos["folio_cck"] = m.group(1).strip() if m else ""

    m_nombres   = re.search(r"Nombre\(s\):\s*\n\s*([A-Z\xc0-\xff ]+)", texto)
    m_apellido1 = re.search(r"Primer apellido:\s*\n\s*([A-Z\xc0-\xff ]+)", texto)
    m_apellido2 = re.search(r"Segundo apellido:\s*\n\s*([A-Z\xc0-\xff ]+)", texto)

    nombres   = limpiar(m_nombres.group(1))   if m_nombres   else ""
    apellido1 = limpiar(m_apellido1.group(1)) if m_apellido1 else ""
    apellido2 = limpiar(m_apellido2.group(1)) if m_apellido2 else ""
    datos["nombre_completo"] = f"{nombres} {apellido1} {apellido2}".strip()

    m = re.search(
        r"Fecha de nacimiento\(dd/mm/aaaa\):.*?\n\s*(\d{2}/\d{2}/\d{4})",
        texto, re.IGNORECASE
    )
    datos["fecha_nacimiento"] = m.group(1).strip() if m else ""

    m = re.search(r"Tel[eé]fono m[oó]vil:.*?\n\s*\d+\s+(\d+)", texto, re.IGNORECASE)
    datos["telefono_movil"] = m.group(1).strip() if m else ""

    m = re.search(
        r"Correo electr[oó]nico:\s*\n\s*[\d\s]+\s+([\w._%+\-]+@[\w.\-]+)",
        texto, re.IGNORECASE
    )
    datos["correo_electronico"] = m.group(1).strip() if m else ""

    m = re.search(
        r"Nombre del vendedor:\s*\n\s*\d+\s+\w+\s+([A-Z\xc0-\xff ]+)",
        texto, re.IGNORECASE
    )
    datos["nombre_vendedor"] = limpiar(m.group(1)) if m else ""



    return datos


def parsear_amortizacion(texto: str) -> dict:
    if not texto:
        return {}

    datos = {}

    m = re.search(r"Folio CCK:\s*(\d+)", texto)
    datos["folio_cck"] = m.group(1).strip() if m else ""

    # Marca — igual que Vehículo: layout de 2 columnas, se corta en el
    # primer salto de 4+ espacios que separa la columna derecha (Tipo de seguro)
    m = re.search(r"^[ \t]*Marca:\s{2,}([^\n]+)", texto, re.IGNORECASE | re.MULTILINE)
    if m:
        datos["marca"] = re.split(r"\s{4,}", m.group(1))[0].strip()
    else:
        datos["marca"] = ""

    m = re.search(r"^[ \t]*Veh[íi]culo:\s{2,}([^\n]+)", texto, re.IGNORECASE | re.MULTILINE)
    datos["vehiculo"] = re.split(r"\s{4,}", m.group(1))[0].strip() if m else ""

    m = re.search(r"Modelo:\s*(\d{4})", texto, re.IGNORECASE)
    datos["modelo_anio"] = m.group(1).strip() if m else ""

    m = re.search(r"Precio de venta \(con IVA\):\s*(\$\s*[\d,]+\.?\d*)", texto, re.IGNORECASE)
    datos["precio_venta_iva"] = extraer_monto(m.group(1)) if m else ""

    m = re.search(r"Enganche\*?:\s*[\d.]+%\s*(\$\s*[\d,]+\.?\d*)", texto, re.IGNORECASE)
    datos["enganche"] = extraer_monto(m.group(1)) if m else ""

    # GAP (monto financiado o de contado, dentro de FINANCIAMIENTO)
    m = re.search(r"GAP:\s*(\$\s*[\d,]+\.?\d*)", texto, re.IGNORECASE)
    gap_raw = m.group(1) if m else ""
    datos["gap_monto"] = extraer_monto(gap_raw) if gap_raw else "$0"
    datos["tiene_gap"] = "SI" if extraer_monto_float(gap_raw) > 0 else "NO"

    m = re.search(r"Plazo:\s*(\d+)\s*meses", texto, re.IGNORECASE)
    datos["plazo_meses"] = m.group(1).strip() if m else ""

    m = re.search(
        r"Tasa de Inter[eé]s \(en t[eé]rminos anuales simples\):\s*([\d.]+%)",
        texto, re.IGNORECASE
    )
    datos["tasa_interes_anual"] = m.group(1).strip() if m else ""

    m = re.search(r"Monto total a financiar:\s*(\$\s*[\d,]+\.?\d*)", texto, re.IGNORECASE)
    datos["monto_total_financiar"] = extraer_monto(m.group(1)) if m else ""

    m_min = re.search(
        r"Mensualidad con seguro primer a[ñn]o \(meses 1-12\):\s*(\$\s*[\d,]+\.?\d*)",
        texto, re.IGNORECASE
    )
    m_max = re.search(
        r"Mensualidad con seguro resto del plazo\s+(\$\s*[\d,]+\.?\d*)",
        texto, re.IGNORECASE
    )

    datos["mensualidad_minima"] = extraer_monto(m_min.group(1)) if m_min else ""
    datos["mensualidad_maxima"] = extraer_monto(m_max.group(1)) if m_max else ""

    # Tipo de seguro (Contado Anual / Multianual Financiado / Multianual Fraccionado)
    m = re.search(r"\bTipo:\s*([^\n]+)", texto, re.IGNORECASE)
    datos["tipo_seguro"] = normalizar_tipo_seguro(m.group(1)) if m else ""

    # Aseguradora (texto libre, ej. INBURSA)
    m = re.search(r"Aseguradora:\s*([^\n]+)", texto, re.IGNORECASE)
    datos["aseguradora"] = limpiar(m.group(1)) if m else ""

    # Accesorios: SI/NO según el monto "Accesorios (10% máximo)"
    m = re.search(
        r"Accesorios\s*\(10%\s*m[aá]ximo\):\s*(\$\s*[\d,]+\.?\d*)",
        texto, re.IGNORECASE
    )
    accesorios_raw = m.group(1) if m else ""
    datos["tiene_accesorios"] = "SI" if extraer_monto_float(accesorios_raw) > 0 else "NO"

    # Garantía extendida: SI/NO según el monto
    m = re.search(r"Garant[ií]a extendida:\s*(\$\s*[\d,]+\.?\d*)", texto, re.IGNORECASE)
    garantia_raw = m.group(1) if m else ""
    datos["tiene_garantia_extendida"] = "SI" if extraer_monto_float(garantia_raw) > 0 else "NO"

    return datos


def procesar_carpeta(carpeta_pdf: str, status: str) -> list:
    archivos = glob.glob(os.path.join(carpeta_pdf, "*.pdf"))
    solicitudes_raw    = [f for f in archivos if "solicitud"    in os.path.basename(f).lower()]
    amortizaciones_raw = [f for f in archivos if "amortizacion" in os.path.basename(f).lower()]

    print(f"\n  ── STATUS: {status} ──")
    print(f"  Solicitudes encontradas   : {len(solicitudes_raw)}")
    print(f"  Amortizaciones encontradas: {len(amortizaciones_raw)}")

    dict_solicitudes = {}
    for ruta in solicitudes_raw:
        datos = parsear_solicitud(extraer_texto_pdf(ruta))
        folio = datos.get("folio_cck", "")
        if folio:
            dict_solicitudes[folio] = datos
            print(f"  ✔ Solicitud  [{folio}] → {datos.get('nombre_completo', 'N/D')}")
        else:
            print(f"  ⚠ No se encontró Folio CCK en: {os.path.basename(ruta)}")

    dict_amortizaciones = {}
    for ruta in amortizaciones_raw:
        datos = parsear_amortizacion(extraer_texto_pdf(ruta))
        folio = datos.get("folio_cck", "")
        if folio:
            dict_amortizaciones[folio] = datos
            print(f"  ✔ Amortización [{folio}] → {datos.get('vehiculo', 'N/D')} {datos.get('modelo_anio', '')}")
        else:
            print(f"  ⚠ No se encontró Folio CCK en: {os.path.basename(ruta)}")

    registros = []
    todos_los_folios = sorted(set(dict_solicitudes) | set(dict_amortizaciones))

    for folio in todos_los_folios:
        sol = dict_solicitudes.get(folio, {})
        amo = dict_amortizaciones.get(folio, {})

        registro = {
            "Folio CCK":               folio,
            "STATUS":                  status,
            "Nombre Completo":         sol.get("nombre_completo", ""),
            "Fecha de Nacimiento":     sol.get("fecha_nacimiento", ""),
            "Teléfono Móvil":          sol.get("telefono_movil", ""),
            "Correo Electrónico":      sol.get("correo_electronico", ""),
            "Nombre del Vendedor":     sol.get("nombre_vendedor", ""),
            "Marca":                   amo.get("marca", ""),
            "Vehículo":                amo.get("vehiculo", ""),
            "Año Modelo":              amo.get("modelo_anio", ""),
            "Accesorios":              amo.get("tiene_accesorios", "NO"),
            "Tipo de Seguro":          amo.get("tipo_seguro", ""),
            "Aseguradora":             amo.get("aseguradora", ""),
            "Precio Venta (c/IVA)":    amo.get("precio_venta_iva", ""),
            "Enganche":                amo.get("enganche", ""),
            "¿Tiene GAP?":            amo.get("tiene_gap", "NO"),
            "Monto GAP":               amo.get("gap_monto", ""),
            "Garantía Extendida":      amo.get("tiene_garantia_extendida", "NO"),
            "Plazo (meses)":           amo.get("plazo_meses", ""),
            "Tasa Interés Anual":      amo.get("tasa_interes_anual", ""),
            "Monto Total a Financiar": amo.get("monto_total_financiar", ""),
            "Mensualidad Mínima":      amo.get("mensualidad_minima", ""),
            "Mensualidad Máxima":      amo.get("mensualidad_maxima", ""),
        }
        registros.append(registro)

        estado_sol = "✔" if folio in dict_solicitudes else "✘ SIN SOLICITUD"
        estado_amo = "✔" if folio in dict_amortizaciones else "✘ SIN AMORTIZACIÓN"
        print(f"  [{folio}] Sol:{estado_sol}  Amor:{estado_amo}  — {registro['Nombre Completo']}")

    return registros



# =======================================================================
# FUNCIÓN PRINCIPAL — punto de entrada para la app (reemplaza a la celda
# "▶️ PROCESAR ZIP DEL MES" del notebook original)
# =======================================================================
def procesar_zip_a_excel(ruta_zip: str, nombre_base: str = None,
                          carpeta_salida: str = ".", limpiar_temporal: bool = True) -> str:
    """
    Recibe la ruta de un ZIP (con la misma estructura de siempre: una
    subcarpeta por STATUS, cada una con sus PDFs de solicitud/amortización)
    y devuelve la ruta del archivo Excel de la bitácora ya generado.

    nombre_base: nombre a usar para el archivo/hoja de salida. Si no se
        indica, se toma del nombre del ZIP (igual que en el notebook).
    carpeta_salida: carpeta donde se guarda el Excel resultante.
    limpiar_temporal: si True, borra los PDFs extraídos al terminar
        (el ZIP de entrada NO se borra, a diferencia del notebook original,
        porque aquí lo maneja Streamlit).
    """
    nombre_base = nombre_base or Path(ruta_zip).stem
    carpeta_trabajo = os.path.join(
        os.path.dirname(carpeta_salida) or ".", f"_extraido_{nombre_base}"
    )
    fecha_proceso = datetime.now().strftime("%Y-%m-%d")
    os.makedirs(carpeta_salida, exist_ok=True)
    archivo_salida = os.path.join(
        carpeta_salida, f"{nombre_base}_{fecha_proceso}_BITACORA.xlsx"
    )

    print(f"\n{'='*60}")
    print(f"   RELACIÓN DE CLIENTES — MG COLIMA ({nombre_base})")
    print(f"{'='*60}")

    carpeta_raiz = extraer_zip(ruta_zip, carpeta_trabajo)
    carpetas_status = detectar_carpetas_status(carpeta_raiz)

    if not carpetas_status:
        raise RuntimeError("No se encontraron subcarpetas de estatus dentro del ZIP.")

    print(f"\n  Carpetas de estatus detectadas: {list(carpetas_status.values())}")

    registros_totales = []
    for nombre_carpeta, status in carpetas_status.items():
        registros_totales.extend(
            procesar_carpeta(os.path.join(carpeta_raiz, nombre_carpeta), status)
        )

    df = pd.DataFrame(registros_totales)

    # --- Exportar a Excel (con formato de encabezado, igual que el original) ---
    from openpyxl.styles import PatternFill, Font, Alignment

    nombre_hoja = nombre_base[:31] if nombre_base else "Reporte"

    with pd.ExcelWriter(archivo_salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nombre_hoja)
        ws = writer.sheets[nombre_hoja]

        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        header_fill = PatternFill("solid", fgColor="CC0000")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"

    if limpiar_temporal:
        shutil.rmtree(carpeta_trabajo, ignore_errors=True)

    print(f"\n{'='*60}")
    print(f"  Reporte generado: {archivo_salida}")
    print(f"  Total de clientes : {len(registros_totales)}")
    print(f"  Fecha de proceso  : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*60}")

    return archivo_salida
